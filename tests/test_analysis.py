import ast
from decimal import Decimal
import io
import json
from dataclasses import FrozenInstanceError
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ldfreq.analysis import (
    AnalysisConfig,
    AnalysisResources,
    SERVER_ONLY_MIN_TOKENS,
    SERVER_ONLY_MIN_TYPES,
    TextDocument,
    analyze_documents,
    analyze_text,
)
from ldfreq.lemmatizers import WordFormLemmatizer
from ldfreq.exporting import payload_to_excel
from ldfreq.privacy import sensitive_paths
from ldfreq.tubelex import TubelexIndex, TubelexRecord
from ldfreq.tokenizer import (
    ASCII_LEGACY_V1,
    DEFAULT_TOKENIZER_POLICY,
)


def _frequency_resources():
    return AnalysisResources(
        lemmatizer=WordFormLemmatizer(),
        rank_map={"common": 1, "word": 1001},
        list_meta={
            "entries": 2,
            "keys": 2,
            "variants": 0,
            "max_rank": 1001,
            "n_levels": 2,
        },
        list_entry={
            "id": "fixture",
            "registry_id": "fixture-open",
            "name": "Open fixture",
            "license": "CC0",
            "loader": lambda _path: None,
        },
        list_path="/operator/private/resources/fixture.csv",
    )


def test_analysis_module_has_no_streamlit_import():
    source = Path("ldfreq/analysis.py").read_text(encoding="utf-8")
    tree = ast.parse(source)
    imported = {
        alias.name.split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.Import)
        for alias in node.names
    }
    imported.update(
        (node.module or "").split(".")[0]
        for node in ast.walk(tree)
        if isinstance(node, ast.ImportFrom)
    )
    assert "streamlit" not in imported


def test_analysis_config_is_typed_frozen_and_validated():
    config = AnalysisConfig(thresholds=[90, 95])
    assert config.thresholds == (90, 95)
    assert isinstance(config.mtld_threshold, float)
    assert config.tokenizer_policy == DEFAULT_TOKENIZER_POLICY
    with pytest.raises(FrozenInstanceError):
        config.min_tokens = 10
    with pytest.raises(ValueError, match="thresholds"):
        AnalysisConfig(thresholds=(0,))
    with pytest.raises(ValueError, match="mattr_window"):
        AnalysisConfig(mattr_window=0)
    for invalid in ("0.72", Decimal("0.72"), True, float("nan"), float("inf")):
        with pytest.raises(ValueError, match="mtld_threshold"):
            AnalysisConfig(mtld_threshold=invalid)
    with pytest.raises(ValueError, match="tokenizer_policy must be one of"):
        AnalysisConfig(tokenizer_policy="Unicode words with apostrophes")


def test_analysis_tokenizer_provenance_drives_actual_tokenization():
    default_result = analyze_text("naïve don’t")
    legacy_result = analyze_text(
        "naïve don’t",
        AnalysisConfig(tokenizer_policy=ASCII_LEGACY_V1),
    )

    assert default_result["n_tokens"] == 2
    assert default_result["payload"]["settings"]["tokenizer_policy"] == (
        DEFAULT_TOKENIZER_POLICY
    )
    assert legacy_result["n_tokens"] == 4
    assert legacy_result["payload"]["settings"]["tokenizer_policy"] == (
        ASCII_LEGACY_V1
    )


def test_analyze_text_returns_aggregate_only_result():
    secret = "quizzacious"
    result = analyze_text(
        f"{secret} {secret} common word",
        AnalysisConfig(thresholds=(90,), msttr_segment=2, mattr_window=2),
        resources=_frequency_resources(),
    )

    assert result["name"] == "Document 001"
    assert result["n_tokens"] == 4
    assert result["n_types"] == 3
    assert result["indices"]["hdd"] is None
    assert result["index_records"]["hdd"]["missing_reason"] == (
        "too_short_for_requested_parameter"
    )
    assert result["index_records"]["hdd"]["requested_parameters"] == {
        "sample_size": 42
    }
    assert result["index_records"]["hdd"]["effective_parameters"] == {}
    assert result["payload"]["panel_a_records"] == result["index_records"]
    assert result["panel_b"]["lfp"][0]["tokens"] == 1
    assert result["payload"]["settings"]["panel_b_mapping_method_id"] == (
        "surface_first_rank_lookup_normalized_fallback_v1"
    )
    assert any(
        "hybrid" in note and "surface-form key" in note
        for note in result["payload"]["method_notes"]
    )
    assert any(
        "not claimed to be numerically comparable to LexTutor" in note
        for note in result["payload"]["method_notes"]
    )
    assert result["list_path"] == "fixture.csv"
    assert result["effective_lemmatizer"] == {"name": "word_form", "version": "-"}
    assert result["payload"]["privacy"] == {
        "source_text_retained": False,
        "source_filename_retained": False,
        "token_level_output_retained": False,
    }
    assert sensitive_paths(result) == []
    assert secret not in repr(result)
    assert "private/resources" not in repr(result)
    assert "<lambda>" not in repr(result)
    json.dumps(result, allow_nan=False)


def test_tubelex_uses_its_treebank_adapter_and_returns_only_aggregates():
    index = TubelexIndex(
        [
            TubelexRecord("do", 40, 8, 6, (40,)),
            TubelexRecord("n't", 30, 7, 5, (30,)),
            TubelexRecord("well-being", 5, 2, 2, (5,)),
        ],
        categories=("education",),
        totals=TubelexRecord("[TOTAL]", 100, 10, 8, (100,)),
        source_vocabulary_size=4,
    )

    result = analyze_text(
        "Don't use well-being or UnseenToken.",
        resources=AnalysisResources(tubelex_index=index),
    )
    tubelex = result["tubelex"]

    assert tubelex["tokens"] == 6
    assert tubelex["covered_tokens"] == 3
    assert tubelex["token_coverage"] == pytest.approx(3 / 6)
    assert tubelex["frequency_zipf_token_mean"] is not None
    assert tubelex["metadata"]["corpus_tokens"] == 100
    assert tubelex["metadata"]["corpus_types"] == 4
    assert result["payload"]["tubelex"] == tubelex
    assert "unseentoken" not in repr(result).lower()


def test_analyze_documents_ignores_names_and_computes_overlap_before_scrubbing():
    progress = []
    response = analyze_documents(
        [
            {"name": "student-a-secret.txt", "text": "alpha beta"},
            {"name": "student-b-secret.txt", "text": "alpha gamma"},
        ],
        AnalysisConfig(msttr_segment=2, mattr_window=2, hdd_sample=2),
        progress=lambda completed, total, label: progress.append(
            (completed, total, label)
        ),
    )

    assert [result["name"] for result in response.results] == [
        "Document 001",
        "Document 002",
    ]
    pair = response.payload["batch_diagnostics"]["overlap_pairs"][0]
    assert pair["shared_types"] == 1
    assert pair["union_types"] == 3
    assert pair["jaccard"] == pytest.approx(1 / 3)
    assert progress == [
        (1, 2, "Document 001"),
        (2, 2, "Document 002"),
    ]
    assert sensitive_paths({"results": response.results, "payload": response.payload}) == []
    representation = repr(response)
    assert "student-a-secret.txt" not in representation
    assert "student-b-secret.txt" not in representation
    assert "alpha" not in representation
    assert "beta" not in representation
    assert "gamma" not in representation


def test_high_entropy_canary_absent_from_exports_and_process_output(capsys):
    canary = "CANARYqzxvplmSECRET"
    filename = "student-CANARY-filename.txt"
    response = analyze_documents(
        [{"name": filename, "text": f"{canary} common word"}],
        resources=_frequency_resources(),
    )

    json_export = json.dumps(response.payload, ensure_ascii=False)
    workbook = load_workbook(
        io.BytesIO(payload_to_excel(response.payload)),
        read_only=True,
        data_only=True,
    )
    excel_cells = "\n".join(
        str(cell)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    captured = capsys.readouterr()

    for output in (repr(response), json_export, excel_cells, captured.out, captured.err):
        assert canary not in output
        assert filename not in output


def test_tokenless_documents_are_skipped_without_echoing_input():
    response = analyze_documents([TextDocument("123 -- ...")])

    assert response.results == ()
    assert response.skipped == (
        {"name": "Document 001", "error": "No tokens found."},
    )
    assert "123 -- ..." not in repr(response)
    with pytest.raises(ValueError, match="No tokens found") as caught:
        analyze_text("123 -- ...")
    assert "123 -- ..." not in str(caught.value)


def test_frequency_resources_are_all_or_nothing():
    with pytest.raises(ValueError, match="supplied together"):
        AnalysisResources(rank_map={"word": 1})


def test_server_only_resource_does_not_expose_server_filename():
    resources = _frequency_resources()
    entry = dict(resources.list_entry)
    entry["delivery_mode"] = "server-side-only"
    entry["source_url"] = "https://example.test/official-list.zip"
    entry["license_url"] = "https://creativecommons.org/licenses/by-sa/4.0/"
    entry["modification_notice"] = "Normalized into a private server index."
    server_only = AnalysisResources(
        lemmatizer=resources.lemmatizer,
        rank_map=resources.rank_map,
        list_meta=resources.list_meta,
        list_entry=entry,
        list_path="/operator/private/resources/secret-list.xlsx",
    )

    words = [
        "common", "word", "alpha", "bravo", "charlie", "delta", "echo",
        "foxtrot", "golf", "hotel", "india", "juliet", "kilo", "lima",
        "mango", "november", "oscar", "papa", "quebec", "romeo",
    ]
    text = " ".join(
        words[index % len(words)] for index in range(SERVER_ONLY_MIN_TOKENS)
    )
    assert len(set(words)) >= SERVER_ONLY_MIN_TYPES
    result = analyze_text(text, resources=server_only)

    assert result["list_path"] is None
    assert result["payload"]["settings"]["list"] is None
    assert result["payload"]["settings"]["list_delivery"] == "server-side-only"
    assert result["payload"]["settings"]["list_source_url"] == entry["source_url"]
    assert result["payload"]["settings"]["list_license_url"] == entry["license_url"]
    assert result["payload"]["settings"]["list_modification_notice"] == entry["modification_notice"]
    assert "secret-list.xlsx" not in repr(result)


def test_server_only_resource_rejects_single_word_oracle_queries():
    resources = _frequency_resources()
    entry = dict(resources.list_entry)
    entry["delivery_mode"] = "server-side-only"
    server_only = AnalysisResources(
        lemmatizer=resources.lemmatizer,
        rank_map=resources.rank_map,
        list_meta=resources.list_meta,
        list_entry=entry,
        list_path="/operator/private/resources/secret-list.xlsx",
    )

    with pytest.raises(ValueError, match="requires at least"):
        analyze_text("common", resources=server_only)
