"""Cross-check the frozen public-v1 contract against executable code.

This test is intentionally strict.  A metric/schema/resource change should fail
until its SemVer impact is reviewed and both v1 scope documents are updated.
"""
from __future__ import annotations

import io
import json
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from ldfreq import OUTPUT_SCHEMA_VERSION, __version__
from ldfreq import batch as BATCH
from ldfreq import exporting as EXPORT
from ldfreq import frequency as FRQ
from ldfreq import indices as IDX
from ldfreq import semantic_network as SEMANTIC
from ldfreq import server_only_gate as SERVER_GATE
from ldfreq import tubelex as TUBELEX
from ldfreq import wordlists as WL
from ldfreq.analysis import (
    AnalysisConfig,
    AnalysisResources,
    analyze_documents,
    analyze_text,
)
from ldfreq.lemmatizers import WordFormLemmatizer


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCOPE_PATH = PROJECT_ROOT / "docs" / "v1-metric-scope.json"
REGISTRY_PATH = PROJECT_ROOT / "data" / "resource_registry.json"
OEWN_MANIFEST_PATH = (
    PROJECT_ROOT
    / "data"
    / "open"
    / "open_english_wordnet"
    / "2025"
    / "manifest.json"
)
TUBELEX_MANIFEST_PATH = (
    PROJECT_ROOT
    / "data"
    / "open"
    / "tubelex"
    / "en"
    / "2025-04-24-7cb5fb36"
    / "manifest.json"
)
GOLDEN_MANIFEST_PATH = (
    PROJECT_ROOT / "tests" / "fixtures" / "v1_golden" / "manifest.json"
)


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _scope() -> dict:
    return _load_json(SCOPE_PATH)


def _resources_with_all_public_metric_sections() -> AnalysisResources:
    tubelex_index = TUBELEX.TubelexIndex(
        [
            TUBELEX.TubelexRecord("alpha", 50, 5, 4, (50,)),
            TUBELEX.TubelexRecord("beta", 30, 4, 3, (30,)),
        ],
        categories=("education",),
        totals=TUBELEX.TubelexRecord("[TOTAL]", 100, 10, 8, (100,)),
        source_vocabulary_size=3,
    )
    return AnalysisResources(
        lemmatizer=WordFormLemmatizer(),
        rank_map={"alpha": 1, "beta": 1001, "gamma": 3001},
        list_meta={
            "entries": 3,
            "keys": 3,
            "variants": 0,
            "max_rank": 3001,
            "n_levels": 4,
            "lookup_unit": "fixture head",
        },
        list_entry={
            "id": "fixture",
            "registry_id": "fixture-green",
            "name": "Frozen-scope fixture",
            "license": "CC0",
            "delivery_mode": "bundled-public",
        },
        list_path="/operator/runtime/fixture.csv",
        semantic_index=SEMANTIC.SemanticNetworkIndex([]),
        tubelex_index=tubelex_index,
    )


def _long_text() -> str:
    return " ".join(["alpha", "beta", "gamma"] * 25)


def _full_payload() -> dict:
    # Schema tests must not depend on whichever NLTK happens to run the test;
    # tokenizer-version enforcement has its own production tests.
    with patch.object(TUBELEX, "tokenize_tubelex_text", side_effect=str.split):
        return analyze_text(
            _long_text(),
            AnalysisConfig(thresholds=(90, 95, 98)),
            resources=_resources_with_all_public_metric_sections(),
        )["payload"]


def _full_batch_payload() -> dict:
    with patch.object(TUBELEX, "tokenize_tubelex_text", side_effect=str.split):
        return analyze_documents(
            [_long_text(), _long_text()],
            AnalysisConfig(thresholds=(90, 95, 98)),
            resources=_resources_with_all_public_metric_sections(),
        ).payload


def test_scope_document_is_valid_and_targets_v1() -> None:
    scope = _scope()
    assert scope["contract_id"] == "ldfreq-public-v1-metric-scope"
    assert scope["contract_version"] == "2.0.0"
    assert scope["target_application_release"] == "1.0.0"
    assert scope["release_line"] == "1.x"
    assert scope["status"] == "frozen"
    assert (PROJECT_ROOT / "docs" / "v1-scope-freeze.md").is_file()


def test_panel_a_keys_directions_and_floors_match_implementation() -> None:
    panel_a = _scope()["metrics"]["panel_a"]
    assert panel_a["keys"] == list(IDX._FUNCS)
    assert panel_a["diversity_direction"] == IDX.DIRECTION
    assert panel_a["project_minimum_tokens"] == IDX.MIN_TOKENS
    assert panel_a["method_ids"] == IDX.METHOD_IDS
    assert panel_a["computational_minimum_tokens_at_default_parameters"] == {
        key: IDX.computational_min_tokens(key) for key in IDX._FUNCS
    }
    records = IDX.all_index_records(["same"] * 10)
    assert list(records) == panel_a["keys"]
    assert list(records["mtld"]) == panel_a["record_keys"]
    assert records["mtld"]["method_id"] == panel_a["method_ids"]["mtld"]
    assert records["mtld"]["requested_parameters"] == records["mtld"][
        "effective_parameters"
    ]
    assert records["mtld"]["advisory_quality_status"] == "below_advisory_floor"
    assert records["mtld"]["status"] == "available"


def test_panel_b_keys_and_conditional_nested_schemas_match_implementation() -> None:
    panel_b_scope = _scope()["metrics"]["panel_b"]
    normalizer = WordFormLemmatizer()
    full = FRQ.panel_b(
        ["alpha", "beta", "gamma"] * 25,
        {"alpha": 1, "beta": 1001, "gamma": 3001},
        normalizer,
        n_levels=4,
    )
    public_full = {
        key: value for key, value in full.items() if not str(key).startswith("_")
    }
    nested = panel_b_scope["nested_keys"]

    assert panel_b_scope["keys"] == list(public_full)
    assert panel_b_scope["mapping_method_id"] == FRQ.PANEL_B_MAPPING_METHOD_ID
    assert list(full["mapping_diagnostics"]) == nested["mapping_diagnostics"]
    assert full["mapping_diagnostics"]["method_id"] == panel_b_scope[
        "mapping_method_id"
    ]
    assert list(full["lfp"][0]) == nested["lfp_row"]
    assert list(full["mean_rank"]) == nested["mean_rank"]
    assert list(full["p_lex"]) == nested["p_lex_complete_segments"]
    assert list(full["s_index"]) == nested["s_index_computable"]
    assert list(full["band_wise"][0]) == nested["band_wise_row"]

    short = FRQ.panel_b(["alpha"], {"alpha": 1}, normalizer, n_levels=1)
    assert list(short["p_lex"]) == nested["p_lex_short_text"]
    assert list(short["s_index"]) == nested["s_index_short_text"]
    assert "_mapped" not in panel_b_scope["keys"]


def test_semantic_and_tubelex_metric_keys_match_runtime_payload() -> None:
    scope_metrics = _scope()["metrics"]
    payload = _full_payload()
    semantic = payload["semantic_network"]
    tubelex = payload["tubelex"]

    assert list(semantic) == scope_metrics["semantic_network"]["output_keys"]
    assert list(tubelex) == scope_metrics["tubelex"]["output_keys"]
    assert list(tubelex["metadata"]) == scope_metrics["tubelex"]["metadata_keys"]

    semantic_measure_keys = set(scope_metrics["semantic_network"]["measure_keys"])
    tubelex_measure_keys = set(scope_metrics["tubelex"]["measure_keys"])
    assert semantic_measure_keys <= set(semantic)
    assert tubelex_measure_keys <= set(tubelex)


def test_single_document_and_batch_json_schemas_match_runtime() -> None:
    json_scope = _scope()["output_contract"]["json"]
    payload = _full_payload()

    assert list(payload) == json_scope["single_document_top_level_keys"]
    assert payload["ldfreq_version"] == __version__
    assert payload["output_schema_version"] == OUTPUT_SCHEMA_VERSION
    assert payload["output_schema_version"] == _scope()["contract_version"]
    assert list(payload["document"]) == json_scope["document_keys"]
    assert list(payload["settings"]) == json_scope["settings_keys"]
    assert payload["settings"]["panel_b_mapping_method_id"] == (
        FRQ.PANEL_B_MAPPING_METHOD_ID
    )
    assert list(payload["privacy"]) == json_scope["privacy_keys"]
    assert payload["panel_a_records"] == {
        key: payload["panel_a_records"][key]
        for key in _scope()["metrics"]["panel_a"]["keys"]
    }
    assert all(
        record["effective_parameters"]
        == (record["requested_parameters"] if record["status"] == "available" else {})
        for record in payload["panel_a_records"].values()
    )
    assert payload["privacy"] == _scope()["output_contract"]["privacy"][
        "payload_values"
    ]

    batch_payload = _full_batch_payload()
    diagnostics = batch_payload["batch_diagnostics"]
    assert list(batch_payload) == json_scope["nonempty_batch_top_level_keys"]
    assert list(batch_payload["batch"]) == json_scope["batch_keys"]
    assert list(diagnostics) == json_scope["batch_diagnostics_keys"]
    for name, expected_keys in json_scope["batch_diagnostic_row_keys"].items():
        assert diagnostics[name], name
        assert list(diagnostics[name][0]) == expected_keys

    skipped_payload = analyze_documents(["123 -- ..."]).payload
    assert list(skipped_payload) == json_scope["all_skipped_batch_top_level_keys"]


def test_serialization_precision_and_golden_fixture_are_frozen() -> None:
    serialization = _scope()["output_contract"]["serialization"]
    manifest = _load_json(GOLDEN_MANIFEST_PATH)

    assert serialization["float_decimal_places"] == EXPORT.EXPORT_FLOAT_DECIMAL_PLACES
    assert serialization["json"]["terminal_newline"] is EXPORT.JSON_TERMINAL_NEWLINE
    assert serialization["xlsx"]["zip_member_timestamp"] == "1980-01-01T00:00:00"
    assert list(EXPORT.XLSX_ZIP_TIMESTAMP) == [1980, 1, 1, 0, 0, 0]
    assert serialization["xlsx"][
        "binary_sha256_normative_only_after_release_image_digest_is_frozen"
    ] is True
    assert manifest["fixture_id"] == "ldfreq-public-v1-golden"
    assert manifest["serialization"]["float_decimal_places"] == (
        EXPORT.EXPORT_FLOAT_DECIMAL_PLACES
    )
    assert manifest["serialization"]["release_image_digest_frozen"] is False


def test_green_runtime_resource_ids_and_runtime_selectors_are_exact() -> None:
    scope = _scope()
    registry = _load_json(REGISTRY_PATH)
    scoped_resources = scope["resources"]["runtime"]
    scoped_ids = {entry["registry_id"] for entry in scoped_resources}
    eligible_registry_entries = {
        entry["id"]: entry
        for entry in registry["resources"]
        if entry["tier"] == "runtime-resource"
        and entry["status"]["level"] == "green"
        and entry["license"]["verified"] is True
        and str(entry["web_use"]["public_saas_processing"]).startswith("allowed")
    }

    assert len(scoped_ids) == len(scoped_resources)
    assert scoped_ids == set(eligible_registry_entries)
    for scoped in scoped_resources:
        registered = eligible_registry_entries[scoped["registry_id"]]
        assert scoped["provisioning"] == registered["provisioning"]["mode"]

    expected_selector_map = {
        entry["runtime_selector_id"]: entry["registry_id"]
        for entry in scoped_resources
        if entry["runtime_selector_id"] is not None
    }
    actual_selector_map = {
        entry["id"]: entry["registry_id"]
        for entry in WL.REGISTRY
        if entry.get("registry_id") in scoped_ids
    }
    assert scope["resources"]["public_panel_b_runtime_selector_ids"] == list(
        expected_selector_map
    )
    assert actual_selector_map == expected_selector_map

    activation = scope["resources"]["server_only_activation"]
    assert activation["default_enabled_ids"] == []
    assert set(activation["eligible_runtime_selector_ids"]) == (
        SERVER_GATE.SERVER_ONLY_ELIGIBLE_IDS
    )
    assert activation["allowlist_with_unknown_id_rejected_as_a_whole"] is True
    assert activation["control_attestation_profile"] == (
        SERVER_GATE.SERVER_ONLY_CONTROL_PROFILE
    )
    assert activation["external_evidence_id_required"] is True
    assert activation["external_evidence_id_is_reference_only"] is True
    assert activation["runtime_verifies_external_controls"] is False

    oewn_manifest = _load_json(OEWN_MANIFEST_PATH)
    tubelex_manifest = _load_json(TUBELEX_MANIFEST_PATH)
    semantic_scope = scope["metrics"]["semantic_network"]
    tubelex_scope = scope["metrics"]["tubelex"]
    assert oewn_manifest["id"] == semantic_scope["artifact_manifest_id"]
    assert tubelex_manifest["id"] == tubelex_scope["artifact_manifest_id"]
    assert TUBELEX.PRODUCTION_RESOURCE_ID == tubelex_scope["registry_resource_id"]


def test_excel_sheet_names_summary_columns_and_detail_rows_are_frozen() -> None:
    excel_scope = _scope()["output_contract"]["excel"]
    payload = _full_payload()
    single_workbook = load_workbook(
        io.BytesIO(EXPORT.payload_to_excel(payload)), read_only=True, data_only=True
    )
    assert single_workbook.sheetnames == excel_scope["always_sheets"]

    summary = EXPORT.summary_rows(payload)
    assert list(summary[0]) == excel_scope["summary_columns"]
    descriptives = EXPORT.descriptive_rows(payload)
    assert [row["measure"] for row in descriptives] == excel_scope[
        "descriptive_measure_keys"
    ]
    assert list(descriptives[0]) == excel_scope["row_columns"]["descriptives"]

    row_builders = {
        "panel_a": EXPORT.panel_a_rows,
        "lfp": EXPORT.lfp_rows,
        "thresholds": EXPORT.coverage_threshold_rows,
        "p_lex_s": EXPORT.p_lex_s_rows,
        "p_lex_dist": EXPORT.p_lex_distribution_rows,
        "s_empirical": EXPORT.s_empirical_rows,
        "band_wise": EXPORT.band_wise_rows,
        "metadata": EXPORT.metadata_rows,
    }
    for sheet, builder in row_builders.items():
        rows = builder(payload)
        assert rows, sheet
        assert list(rows[0]) == excel_scope["row_columns"][sheet]

    semantic_rows = EXPORT.semantic_network_rows(payload)
    tubelex_rows = EXPORT.tubelex_rows(payload)
    assert list(semantic_rows[0]) == [
        "document",
        *_scope()["metrics"]["semantic_network"]["output_keys"],
    ]
    assert list(tubelex_rows[0]) == [
        "document",
        *_scope()["metrics"]["tubelex"]["output_keys"],
    ]

    batch_payload = _full_batch_payload()
    batch_workbook = load_workbook(
        io.BytesIO(EXPORT.payload_to_excel(batch_payload)),
        read_only=True,
        data_only=True,
    )
    assert batch_workbook.sheetnames == [
        *excel_scope["always_sheets"],
        *excel_scope["batch_only_sheets"],
    ]
    off_list_values = [
        value
        for row in batch_workbook["off_list"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    assert off_list_values == []


def test_batch_row_contract_matches_batch_module() -> None:
    expected = _scope()["output_contract"]["json"]["batch_diagnostic_row_keys"]
    results = [
        {
            "name": "Document 001",
            "n_tokens": 2,
            "indices": {key: 1.0 for key in IDX._FUNCS},
            "a_tokens": ["alpha", "beta"],
            "panel_b": {
                "lfp": [
                    {
                        "level": "K1",
                        "tokens": 2,
                        "types": 2,
                        "coverage_%": 100.0,
                        "cumulative_%": 100.0,
                    }
                ]
            },
        },
        {
            "name": "Document 002",
            "n_tokens": 2,
            "indices": {key: 1.0 for key in IDX._FUNCS},
            "a_tokens": ["alpha", "gamma"],
            "panel_b": {"lfp": []},
        },
    ]
    for result in results:
        records = IDX.all_index_records(result["a_tokens"])
        result["indices"] = {
            key: record["value"] for key, record in records.items()
        }
        result["index_records"] = records
    produced = {
        "bands": BATCH.band_rows(results),
        "reliability": BATCH.reliability_rows(results),
        "overlap_matrix": BATCH.overlap_matrix_rows(results),
        "overlap_pairs": BATCH.overlap_pair_rows(results),
    }
    for name, rows in produced.items():
        assert rows, name
        assert list(rows[0]) == expected[name]


def test_unimplemented_and_predictive_outputs_are_not_in_metric_scope() -> None:
    scope = _scope()
    metric_ids = {
        *scope["metrics"]["panel_a"]["keys"],
        *scope["metrics"]["panel_b"]["keys"],
        *scope["metrics"]["semantic_network"]["measure_keys"],
        *scope["metrics"]["tubelex"]["measure_keys"],
    }
    prohibited_fragments = (
        "ellipse",
        "cefr",
        "grade",
        "proficiency",
        "authorship",
        "category_entropy",
        "masc",
        "oanc",
    )
    assert not any(
        fragment in metric_id.lower()
        for metric_id in metric_ids
        for fragment in prohibited_fragments
    )
    non_goal_ids = {item["id"] for item in scope["non_goals"]}
    assert "ellipse_rating_prediction" in non_goal_ids
    assert scope["metrics"]["tubelex"]["category_entropy_included"] is False
    assert scope["metrics"]["semantic_network"]["word_sense_disambiguation"] is False

    scoped_resource_ids = {
        item["registry_id"] for item in scope["resources"]["runtime"]
    }
    assert not any("ellipse" in resource_id for resource_id in scoped_resource_ids)


def test_semver_policy_is_conservative_about_breaking_numerical_changes() -> None:
    semver = _scope()["semver"]
    assert any("formula" in rule for rule in semver["major"]["required_for"])
    assert any("resource identity" in rule for rule in semver["major"]["required_for"])
    assert any("privacy" in rule for rule in semver["major"]["required_for"])
    assert "golden fixtures" in semver["patch"]["numerical_bugfix_rule"]
