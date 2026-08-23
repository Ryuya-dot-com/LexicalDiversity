import copy
import io
import json
import math
import unittest
import zipfile

from openpyxl import load_workbook

from ldfreq import exporting
from ldfreq import indices as IDX


class ExportingTests(unittest.TestCase):
    def test_panel_a_rows_expose_structured_method_contract(self):
        records = IDX.all_index_records(["alpha", "beta", "alpha"])
        payload = {
            "document": {"name": "Document 001"},
            "n_tokens": 3,
            "settings": {
                "msttr_segment": 50,
                "mattr_window": 50,
                "hdd_sample": 42,
                "mtld_threshold": 0.72,
                "vocd_seed": 42,
            },
            "panel_a": {key: record["value"] for key, record in records.items()},
            "panel_a_records": records,
        }

        rows = exporting.panel_a_rows(payload)
        mattr = next(row for row in rows if row["index_key"] == "mattr")

        self.assertEqual(mattr["status"], "missing")
        self.assertEqual(
            mattr["missing_reason"],
            "too_short_for_requested_parameter",
        )
        self.assertEqual(mattr["method_id"], IDX.METHOD_IDS["mattr"])
        self.assertEqual(
            json.loads(mattr["requested_parameters"]),
            {"window_length": 50},
        )
        self.assertEqual(json.loads(mattr["effective_parameters"]), {})

    def test_panel_a_rows_rejects_absent_or_inconsistent_records(self):
        records = IDX.all_index_records(["alpha", "beta", "alpha"])
        payload = {
            "document": {"name": "Document 001"},
            "n_tokens": 3,
            "panel_a": {key: record["value"] for key, record in records.items()},
            "panel_a_records": records,
        }

        without_records = copy.deepcopy(payload)
        del without_records["panel_a_records"]
        with self.assertRaisesRegex(ValueError, "Panel A schema 2.0"):
            exporting.panel_a_rows(without_records)

        incomplete = copy.deepcopy(payload)
        del incomplete["panel_a_records"]["mattr"]
        with self.assertRaisesRegex(ValueError, "Panel A schema 2.0"):
            exporting.panel_a_rows(incomplete)

        inconsistent = copy.deepcopy(payload)
        inconsistent["panel_a_records"]["mattr"]["effective_parameters"] = {
            "window_length": 50
        }
        with self.assertRaisesRegex(ValueError, "Panel A schema 2.0"):
            exporting.panel_a_rows(inconsistent)

        unknown_reason = copy.deepcopy(payload)
        unknown_reason["panel_a_records"]["mattr"]["missing_reason"] = "unknown"
        with self.assertRaisesRegex(ValueError, "Panel A schema 2.0"):
            exporting.panel_a_rows(unknown_reason)

    def test_payload_to_json_uses_frozen_precision_and_terminal_newline(self):
        payload = {
            "value": 1.1234567890129,
            "negative_zero": -0.0,
            "missing": math.nan,
            "unicode": "語彙",
        }

        serialized = exporting.payload_to_json(payload)

        self.assertTrue(serialized.endswith("\n"))
        self.assertNotIn("NaN", serialized)
        self.assertNotIn("\\u8a9e", serialized)
        self.assertEqual(
            json.loads(serialized),
            {
                "value": 1.123456789013,
                "negative_zero": 0.0,
                "missing": None,
                "unicode": "語彙",
            },
        )

    def test_public_export_rejects_infinity(self):
        with self.assertRaisesRegex(ValueError, "infinite"):
            exporting.payload_to_json({"value": math.inf})

    def test_payload_to_excel_writes_batch_summary_and_detail_sheets(self):
        records_a = IDX.all_index_records(["alpha", "alpha", "beta"])
        records_b = IDX.all_index_records(["alpha", "beta"])
        payload = {
            "ldfreq_version": "test",
            "batch_diagnostics": {
                "bands": [{"document": "a.txt", "level": "K1", "coverage_%": 66.67}],
                "reliability": [{"document": "a.txt", "index": "TTR", "status": "available"}],
                "off_list": [{"document": "a.txt", "head": "xray", "count": 1}],
                "overlap_matrix": [{"document_a": "a.txt", "document_b": "b.txt", "jaccard": 0.5}],
                "overlap_pairs": [{"document_a": "a.txt", "document_b": "b.txt", "jaccard": 0.5}],
            },
            "documents": [
                {
                    "document": {"name": "a.txt"},
                    "settings": {
                        "unit": "flemma",
                        "lemmatizer": "word_form",
                        "lemmatizer_version": "-",
                        "list_name": "Test List",
                    },
                    "method_notes": ["Tokenizer policy: test"],
                    "n_tokens": 3,
                    "n_types": 2,
                    "panel_a": {
                        key: record["value"] for key, record in records_a.items()
                    },
                    "panel_a_records": records_a,
                    "panel_b": {
                        "lfp": [{"level": "K1", "tokens": 2, "types": 1, "coverage_%": 66.67, "cumulative_%": 66.67}],
                        "coverage_threshold": {90: None},
                        "advanced_guiraud": 0.58,
                        "pct_beyond_k": 50.0,
                        "mean_rank": {"pct_off_list": 33.33},
                        "p_lex": {"lambda": None, "n_segments": 0},
                        "s_index": {"S": None, "capped": None},
                        "band_wise": [{"level": "K1", "tokens": 2, "types": 1, "Min N": 50}],
                    },
                    "tubelex": {
                        "token_coverage": 2 / 3,
                        "type_coverage": 0.5,
                        "frequency_zipf_token_mean": 4.2,
                        "frequency_zipf_type_mean": 3.8,
                        "video_log10_prevalence_token_mean": -2.1,
                        "video_log10_prevalence_type_mean": -2.4,
                        "channel_log10_prevalence_token_mean": -2.3,
                        "channel_log10_prevalence_type_mean": -2.6,
                    },
                },
                {
                    "document": {"name": "b.txt"},
                    "settings": {"unit": "token", "list_name": "Test List"},
                    "n_tokens": 2,
                    "n_types": 2,
                    "panel_a": {
                        key: record["value"] for key, record in records_b.items()
                    },
                    "panel_a_records": records_b,
                    "panel_b": None,
                },
            ],
        }

        workbook = load_workbook(io.BytesIO(exporting.payload_to_excel(payload)), read_only=True)

        self.assertIn("summary", workbook.sheetnames)
        self.assertIn("descriptives", workbook.sheetnames)
        self.assertIn("panel_a", workbook.sheetnames)
        self.assertIn("metadata", workbook.sheetnames)
        self.assertIn("reliability", workbook.sheetnames)
        self.assertIn("off_list", workbook.sheetnames)
        self.assertIn("overlap_pairs", workbook.sheetnames)
        self.assertIn("semantic_network", workbook.sheetnames)
        self.assertIn("tubelex", workbook.sheetnames)
        summary_rows = list(workbook["summary"].iter_rows(values_only=True))
        self.assertEqual(summary_rows[0][0], "document")
        self.assertEqual(summary_rows[1][0], "a.txt")
        self.assertEqual(summary_rows[2][0], "b.txt")

        descriptives = exporting.descriptive_rows(payload)
        ttr = next(row for row in descriptives if row["measure"] == "ttr")
        self.assertEqual(ttr["n"], 2)
        self.assertEqual(ttr["missing"], 0)
        self.assertAlmostEqual(ttr["mean"], 5 / 6, places=12)
        mattr = next(row for row in descriptives if row["measure"] == "mattr")
        self.assertEqual(mattr["n"], 0)
        self.assertEqual(mattr["missing"], 2)

    def test_payload_to_excel_is_byte_deterministic_and_has_fixed_zip_metadata(self):
        records = IDX.all_index_records(["alpha"])
        payload = {
            "ldfreq_version": "test",
            "document": {"name": "Document 001"},
            "settings": {},
            "method_notes": [],
            "privacy": {},
            "n_tokens": 1,
            "n_types": 1,
            "panel_a": {key: record["value"] for key, record in records.items()},
            "panel_a_records": records,
            "panel_b": None,
            "semantic_network": None,
            "tubelex": None,
        }

        first = exporting.payload_to_excel(payload)
        second = exporting.payload_to_excel(payload)

        self.assertEqual(first, second)
        with zipfile.ZipFile(io.BytesIO(first)) as workbook_zip:
            self.assertTrue(workbook_zip.infolist())
            self.assertTrue(
                all(
                    member.date_time == exporting.XLSX_ZIP_TIMESTAMP
                    for member in workbook_zip.infolist()
                )
            )
            core = workbook_zip.read("docProps/core.xml").decode("utf-8")
        self.assertIn("1980-01-01T00:00:00Z", core)
        self.assertIn("<dc:creator", core)
        self.assertIn(">ldfreq</dc:creator>", core)


if __name__ == "__main__":
    unittest.main()
