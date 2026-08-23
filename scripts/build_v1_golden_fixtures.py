#!/usr/bin/env python3
"""Build or verify the public-v1 canonical analysis/export fixtures offline."""
from __future__ import annotations

import argparse
import hashlib
import json
import platform
import sys
from dataclasses import asdict
from importlib import metadata
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from ldfreq import (
    OUTPUT_SCHEMA_VERSION,
    RELEASE_PHASE,
    TARGET_APPLICATION_RELEASE,
    __version__,
)
from ldfreq.analysis import AnalysisConfig
from ldfreq.exporting import (
    EXPORT_FLOAT_DECIMAL_PLACES,
    canonical_export_value,
    payload_to_excel,
    payload_to_json,
)
from ldfreq.isolated import (
    IsolationLimits,
    ResourceSpec,
    analyze_documents_isolated,
)
from scripts.check_runtime_environment import (
    PRODUCTION_REQUIREMENTS,
    PRODUCTION_WATCHDOG_WHEEL_LOCK,
    PRODUCTION_WHEEL_LOCK,
    read_exact_pins,
    runtime_environment_violations,
)


FIXTURE_ROOT = PROJECT_ROOT / "tests" / "fixtures" / "v1_golden"
INPUT_PATHS = (
    FIXTURE_ROOT / "document-001.txt",
    FIXTURE_ROOT / "document-002.txt",
)
EXPECTED_PATHS = {
    "single_json": FIXTURE_ROOT / "expected-single.json",
    "batch_json": FIXTURE_ROOT / "expected-batch.json",
    "single_workbook": FIXTURE_ROOT / "expected-single-workbook.json",
    "batch_workbook": FIXTURE_ROOT / "expected-batch-workbook.json",
}
MANIFEST_PATH = FIXTURE_ROOT / "manifest.json"
BASE_IMAGE_IDENTITY_PATH = PROJECT_ROOT / "deploy" / "cloud-run" / "base-image.json"


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json_bytes(value: dict[str, Any]) -> bytes:
    return payload_to_json(value).encode("utf-8")


def workbook_snapshot(workbook_bytes: bytes) -> dict[str, Any]:
    """Return the semantic sheet/cell representation used across platforms."""

    workbook = load_workbook(
        BytesIO(workbook_bytes),
        read_only=True,
        data_only=False,
    )
    sheets: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            max_row = int(worksheet.max_row or 1)
            max_column = int(worksheet.max_column or 1)
            rows = [
                canonical_export_value(list(row))
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_column,
                    values_only=True,
                )
            ]
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": max_row,
                    "max_column": max_column,
                    "rows": rows,
                }
            )
    finally:
        workbook.close()
    return {"sheets": sheets}


def _require_clean_runtime() -> dict[str, str]:
    violations, pins = runtime_environment_violations()
    if violations:
        details = "\n".join(f"- {violation}" for violation in violations)
        raise RuntimeError(f"Golden fixture runtime contract is blocked:\n{details}")
    return pins


def build_artifacts() -> tuple[dict[str, bytes], dict[str, Any]]:
    """Run one canonical batch and return expected files plus manifest data."""

    pins = _require_clean_runtime()
    texts = [path.read_text(encoding="utf-8") for path in INPUT_PATHS]
    config = AnalysisConfig()
    response = analyze_documents_isolated(
        texts,
        config,
        ResourceSpec(
            list_id="ngsl",
            lemmatizer_name="open_flemma",
            semantic_enabled=True,
            tubelex_enabled=True,
        ),
        limits=IsolationLimits(deadline_seconds=120.0),
    )
    if len(response.payload.get("documents") or []) != 2:
        raise RuntimeError("Canonical batch did not return exactly two documents")

    batch_payload = response.payload
    single_payload = batch_payload["documents"][0]
    # The application builds Excel from the already serialized JSON string.
    # Reproduce that exact boundary so both download formats share one numeric
    # precision contract.
    single_json = canonical_json_bytes(single_payload)
    batch_json = canonical_json_bytes(batch_payload)
    single_export_payload = json.loads(single_json)
    batch_export_payload = json.loads(batch_json)
    single_xlsx = payload_to_excel(single_export_payload)
    batch_xlsx = payload_to_excel(batch_export_payload)
    files = {
        "single_json": single_json,
        "batch_json": batch_json,
        "single_workbook": canonical_json_bytes(workbook_snapshot(single_xlsx)),
        "batch_workbook": canonical_json_bytes(workbook_snapshot(batch_xlsx)),
    }

    generator = {
        "python": platform.python_version(),
        "implementation": platform.python_implementation(),
        "platform": platform.platform(),
        "production_version_graph_sha256": sha256_file(PRODUCTION_REQUIREMENTS),
        "production_wheel_lock_sha256": sha256_file(PRODUCTION_WHEEL_LOCK),
        "production_watchdog_wheel_lock_sha256": sha256_file(
            PRODUCTION_WATCHDOG_WHEEL_LOCK
        ),
        "base_image_identity_sha256": sha256_file(BASE_IMAGE_IDENTITY_PATH),
        "installed_versions": {
            name: metadata.version(name) for name in sorted(pins)
        },
    }
    manifest = {
        "fixture_id": "ldfreq-public-v1-golden",
        "schema_version": 3,
        "status": "canonical-pre-v1",
        "license": "CC0-1.0",
        "human_or_learner_writing": False,
        "external_api_calls": 0,
        "release_identity": {
            "application_version": __version__,
            "output_schema_version": OUTPUT_SCHEMA_VERSION,
            "target_application_release": TARGET_APPLICATION_RELEASE,
            "release_phase": RELEASE_PHASE,
        },
        "analysis": {
            "documents": 2,
            "config": canonical_export_value(asdict(config)),
            "resources": {
                "list_id": "ngsl",
                "lemmatizer_name": "open_flemma",
                "semantic_enabled": True,
                "tubelex_enabled": True,
            },
        },
        "serialization": {
            "float_decimal_places": EXPORT_FLOAT_DECIMAL_PLACES,
            "json_utf8": True,
            "json_terminal_newline": True,
            "xlsx_semantic_snapshot_normative_across_platforms": True,
            "xlsx_binary_hash_normative_only_in_frozen_release_image": True,
            "release_image_digest_frozen": False,
        },
        "inputs": {
            path.name: {
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in INPUT_PATHS
        },
        "expected_files": {
            EXPECTED_PATHS[key].name: {
                "bytes": len(payload),
                "sha256": sha256_bytes(payload),
            }
            for key, payload in files.items()
        },
        "provisional_xlsx_binary": {
            "single": {
                "bytes": len(single_xlsx),
                "sha256": sha256_bytes(single_xlsx),
            },
            "batch": {
                "bytes": len(batch_xlsx),
                "sha256": sha256_bytes(batch_xlsx),
            },
        },
        "generator": generator,
    }
    return files, manifest


def write_artifacts(files: dict[str, bytes], manifest: dict[str, Any]) -> None:
    for key, payload in files.items():
        EXPECTED_PATHS[key].write_bytes(payload)
    MANIFEST_PATH.write_bytes(canonical_json_bytes(manifest))


def check_artifacts(files: dict[str, bytes], manifest: dict[str, Any]) -> list[str]:
    violations: list[str] = []
    for key, expected in files.items():
        path = EXPECTED_PATHS[key]
        if not path.is_file():
            violations.append(f"missing expected file: {path.relative_to(PROJECT_ROOT)}")
        elif path.read_bytes() != expected:
            violations.append(f"golden output differs: {path.relative_to(PROJECT_ROOT)}")
    if not MANIFEST_PATH.is_file():
        violations.append(f"missing manifest: {MANIFEST_PATH.relative_to(PROJECT_ROOT)}")
    else:
        recorded = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
        # Platform and exact patch provenance may differ while the normative
        # JSON and workbook snapshots remain equal. Do not require generator
        # identity equality until the release image digest is frozen.
        for key in (
            "fixture_id",
            "schema_version",
            "status",
            "license",
            "human_or_learner_writing",
            "external_api_calls",
            "release_identity",
            "analysis",
            "serialization",
            "inputs",
            "expected_files",
        ):
            if recorded.get(key) != manifest.get(key):
                violations.append(f"manifest field differs: {key}")
    return violations


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--write", action="store_true", help="replace expected fixtures")
    mode.add_argument("--check", action="store_true", help="recompute without writing")
    args = parser.parse_args()

    try:
        files, manifest = build_artifacts()
    except Exception as exc:
        print(f"v1 golden fixtures: BLOCKED ({type(exc).__name__}: {exc})", file=sys.stderr)
        return 1
    if args.write:
        write_artifacts(files, manifest)
        print("v1 golden fixtures: WROTE canonical outputs")
        return 0

    violations = check_artifacts(files, manifest)
    if violations:
        print("v1 golden fixtures: BLOCKED", file=sys.stderr)
        for violation in violations:
            print(f"- {violation}", file=sys.stderr)
        return 1
    print("v1 golden fixtures: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
