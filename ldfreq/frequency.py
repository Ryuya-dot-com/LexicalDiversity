"""Panel B: frequency-list–based richness / sophistication measures.

Everything here depends on a ranked word list (e.g. New JACET8000). A list is
loaded into either a ``form/variant -> rank`` dict or a richer
``form/variant -> {head, rank}`` dict; ``level = (rank-1)//1000 + 1``.

Measures
--------
* Lexical Frequency Profile (LFP)  — token coverage per band + cumulative.
* Coverage thresholds              — selected-list band reaching 90/95/98 %.
* Advanced Guiraud (AG)            — advanced_types / sqrt(N).
* % beyond-K                       — share of distinct word families beyond band K (off-list incl.).
* Mean rank / mean log-rank        — central tendency of frequency.
* P_Lex (lambda)                   — Meara & Bell (2001): hard words per 10-word segment, Poisson curve-fit.
* S (Kojima & Yamashita 2014)      — Eq.4 coverage-curve fit C(x)=ln(x)/ln(S)*100.
* Band-wise diversity              — per-band token/type counts and short-band warnings.

Caveat: coverage here is selected-list matched coverage. Proper nouns, marginal
words, acronyms, and other potentially known items are not automatically credited
unless they match the selected list/normalizer. P_Lex and S use the selected list's
ranks, not Kojima & Yamashita's BNC-spoken *family* lists. S in particular assumes
coverage approaches 100 % within the sampled ranks; with incomplete selected-list
coverage it often does not, so S reports ``capped`` and is then not numerically
interpretable.
"""
from __future__ import annotations

import csv
import glob
import math
import os
import re
from collections import Counter, defaultdict

from . import indices as IDX

OFF_LIST = None  # sentinel rank for tokens not found in the list
PANEL_B_MAPPING_METHOD_ID = "surface_first_rank_lookup_normalized_fallback_v1"


# --------------------------------------------------------------------------- #
# Word-list loading
# --------------------------------------------------------------------------- #
def load_ranked_list(path: str, *, encoding="utf-8-sig"):
    """Load a 2-column ``rank,word`` CSV (New JACET8000 layout).

    Handles:
      * parenthetical variants  -> "mom (mommy, mum, …)" maps mom/mommy/mum/… to one rank
      * spreadsheet artifacts   -> "TRUE"/"FALSE" become "true"/"false" via lower-casing
    Returns ``(rank_map, meta)`` where rank_map: head/variant(lower) -> int rank.
    """
    rank: dict[str, int] = {}
    n_variants = 0
    max_rank = 0
    with open(path, encoding=encoding, newline="") as fh:
        reader = csv.reader(fh)
        for row in reader:
            if not row or len(row) < 2:
                continue
            if not row[0].strip().isdigit():  # header line ("NJ8,Word")
                continue
            r = int(row[0])
            raw = row[1].strip()
            head = re.sub(r"\s*\([^)]*\)\s*$", "", raw).lower()
            if head:
                rank.setdefault(head, r)
            m = re.search(r"\(([^)]*)\)", raw)
            if m:
                for v in re.split(r"[,\s]+", m.group(1)):
                    v = v.strip().lower()
                    if v:
                        rank.setdefault(v, r)
                        n_variants += 1
            max_rank = max(max_rank, r)
    meta = {"entries": len({v for v in rank.values()}), "keys": len(rank),
            "variants": n_variants, "max_rank": max_rank,
            "lookup_unit": "listed_surface_form_or_normalized_fallback",
            "n_levels": (max_rank + 999) // 1000}
    return rank, meta


def load_ngsl(dir_path: str, *, encoding="utf-8-sig"):
    """Load the NGSL files shipped by the New General Service List project.

    The local download is a directory rather than a two-column ``rank,word`` CSV.
    ``NGSL_1.2_stats.csv`` supplies the frequency rank and
    ``NGSL_1.2_lemmatized_for_research.csv`` supplies forms belonging to each
    lemma. If the lemmatized file is absent, the loader still returns headword
    ranks from the stats file.
    """
    stats_path = dir_path
    base_dir = os.path.dirname(dir_path) or "."
    if os.path.isdir(dir_path):
        base_dir = dir_path
        stats_path = os.path.join(dir_path, "NGSL_1.2_stats.csv")

    lemma_rank: dict[str, int] = {}
    with open(stats_path, encoding=encoding, newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            lemma = (row.get("Lemma") or "").strip().lower()
            rank_raw = (row.get("SFI Rank") or "").strip()
            if lemma and rank_raw.isdigit():
                lemma_rank.setdefault(lemma, int(rank_raw))

    rank: dict[str, int] = dict(lemma_rank)
    n_variants = 0
    forms_path = os.path.join(base_dir, "NGSL_1.2_lemmatized_for_research.csv")
    if os.path.exists(forms_path):
        with open(forms_path, encoding=encoding, newline="") as fh:
            for row in csv.reader(fh):
                if not row or row[0].startswith("##"):
                    continue
                head = row[0].strip().lower()
                r = lemma_rank.get(head)
                if r is None:
                    continue
                for form in row:
                    form = form.strip().lower()
                    if not form:
                        continue
                    if form != head:
                        n_variants += 1
                    rank.setdefault(form, r)

    max_rank = max(lemma_rank.values(), default=0)
    meta = {"entries": len(lemma_rank), "keys": len(rank),
            "variants": n_variants, "max_rank": max_rank,
            "lookup_unit": "listed_surface_form_at_lemma_rank_or_normalized_fallback",
            "n_levels": (max_rank + 999) // 1000}
    return rank, meta


_ORD_RE = re.compile(r"(\d+)\s*(?:st|nd|rd|th)\b", re.IGNORECASE)
_BNC_COCA_LEVEL_RE = re.compile(r"^\s*(\d+)\s*k\s*$", re.IGNORECASE)
_BNC_COCA_RELATED_FORM_RE = re.compile(r"([^,()]+?)\s*\((\d+)\)")
_RANGE_FILE_ORDER_RE = re.compile(r"(?:basewrd|baseword|^)(\d+)|^(\d+)[_\-\s]", re.IGNORECASE)
_LIST_WORD_RE = re.compile(r"[A-Za-z][A-Za-z']*")


def _cell_text(value) -> str:
    return "" if value is None else str(value)


def load_headword_bands(dir_path, *, encoding="utf-8-sig"):
    """Load Nation's BNC/COCA *headword* lists from a directory.

    Files named like ``headwords 1st 1000.txt`` … ``headwords 10th 1000.txt`` each hold
    1000 headwords (one per line). Band K = the ordinal in the filename; rank within a
    band = line order, so global rank = (K-1)*1000 + line.

    NOTE: these are HEADWORDS only — the word-family *members* (derivations) are NOT
    included (they ship with AntWordProfiler/Range). Profiling therefore uses hybrid
    headword-entry coverage: direct listed headwords plus normalized misses that
    resolve to a listed headword. Derivations not reducible by the normalizer remain
    off-list. True word-family coverage needs the member (basewrd) lists.
    """
    def band_of(fp):
        m = _ORD_RE.search(os.path.basename(fp))
        return int(m.group(1)) if m else 9999

    def read_lines(fp):
        try:
            with open(fp, encoding=encoding) as fh:
                return fh.readlines()
        except UnicodeDecodeError:  # some files are Latin-1, not UTF-8
            with open(fp, encoding="latin-1") as fh:
                return fh.readlines()

    files = glob.glob(os.path.join(dir_path, "headwords*1000.txt"))
    if not files:
        files = glob.glob(os.path.join(dir_path, "**", "headwords*1000.txt"), recursive=True)
    files = sorted(files, key=band_of)
    rank, max_band = {}, 0
    for fp in files:
        b = band_of(fp)
        if b == 9999:
            continue
        max_band = max(max_band, b)
        for i, line in enumerate(read_lines(fp)):
            w = line.strip().lower()
            if w and w not in rank:
                rank[w] = (b - 1) * 1000 + i + 1
    meta = {"entries": len(rank), "keys": len(rank), "variants": 0,
            "max_rank": max_band * 1000, "n_levels": max_band,
            "lookup_unit": "listed_headword_or_normalized_fallback"}
    return rank, meta


def load_bnc_coca_families_xlsx(path: str):
    """Load Nation's BNC/COCA word-family spreadsheet.

    Expected columns are ``List``, ``Headword``, ``Related forms``, and
    ``Total frequency``. Each related form maps to the family head and frequency
    band. The official spreadsheet has a few bands with 999/1001/1002 rows, so
    rank values are clamped within the displayed band to keep K-level assignment
    stable.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [_cell_text(value).strip().lower() for value in next(rows, ())]
        try:
            list_idx = next(i for i, value in enumerate(header) if value.startswith("list"))
            head_idx = next(i for i, value in enumerate(header) if value.startswith("headword"))
            related_idx = next(i for i, value in enumerate(header) if value.startswith("related"))
            freq_idx = next(i for i, value in enumerate(header) if value.startswith("total"))
        except StopIteration as exc:
            raise ValueError(f"Unexpected BNC/COCA family-list header in {path}") from exc

        rank: dict[str, dict[str, int | str]] = {}
        entries = 0
        collisions = 0
        max_level = 0
        per_level_seen: Counter[int] = Counter()

        for row in rows:
            label = _cell_text(row[list_idx]).strip()
            match = _BNC_COCA_LEVEL_RE.match(label)
            if not match:
                continue
            level = int(match.group(1))
            max_level = max(max_level, level)
            per_level_seen[level] += 1
            ordinal = min(per_level_seen[level], 1000)
            family_rank = (level - 1) * 1000 + ordinal
            family_head = _cell_text(row[head_idx]).strip().lower()
            if not family_head:
                continue
            total_frequency = row[freq_idx] or 0
            entries += 1
            forms = {
                form.strip().lower()
                for form, _freq in _BNC_COCA_RELATED_FORM_RE.findall(_cell_text(row[related_idx]))
                if form.strip()
            }
            forms.add(family_head)
            for form in forms:
                entry = {
                    "head": family_head,
                    "rank": family_rank,
                    "level": level,
                    "frequency": int(total_frequency) if isinstance(total_frequency, (int, float)) else 0,
                }
                if form in rank:
                    collisions += 1
                    continue
                rank[form] = entry

        meta = {
            "entries": entries,
            "keys": len(rank),
            "variants": max(0, len(rank) - entries),
            "collisions": collisions,
            "max_rank": max_level * 1000,
            "n_levels": max_level,
            "lookup_unit": "word_family",
        }
        return rank, meta
    finally:
        wb.close()


def _read_text_lines(path: str, *, encoding="utf-8-sig") -> list[str]:
    try:
        with open(path, encoding=encoding) as fh:
            return fh.readlines()
    except UnicodeDecodeError:
        with open(path, encoding="latin-1") as fh:
            return fh.readlines()


def _range_file_order(path: str) -> tuple[int, str]:
    name = os.path.basename(path)
    match = _RANGE_FILE_ORDER_RE.search(name)
    if match:
        return int(next(group for group in match.groups() if group)), name.lower()
    return 9999, name.lower()


def _range_level_files(path: str) -> list[str]:
    if os.path.isfile(path):
        return [path]
    patterns = [
        "basewrd*",
        "baseword*",
        "[0-9]*.txt",
        "*.txt",
    ]
    files: list[str] = []
    for pattern in patterns:
        files.extend(glob.glob(os.path.join(path, pattern)))
    unique = sorted({fp for fp in files if os.path.isfile(fp)}, key=_range_file_order)
    return unique


def _first_list_word(line: str) -> str | None:
    match = _LIST_WORD_RE.search(line)
    return match.group(0).lower() if match else None


def load_range_baseword_lists(path: str, *, encoding="utf-8-sig"):
    """Load AntWordProfiler/Range-style baseword level-list files.

    A directory is treated as multiple levels, ordered by filenames such as
    ``BASEWRD1.txt``, ``BASEWRD2.txt``, or ``1_gsl_1st_1000.txt``. An unindented
    word starts a new family; an indented word is a member of the current family.
    Plain one-word-per-line lists are also accepted, with each word treated as
    its own family. Blank lines and lines without a word token are ignored.
    """
    files = _range_level_files(path)
    if not files:
        raise FileNotFoundError(f"No AntWordProfiler/Range level-list files found at: {path}")

    rank: dict[str, dict[str, int | str]] = {}
    entries = 0
    variants = 0
    collisions = 0
    max_level = 0

    for level, fp in enumerate(files, start=1):
        max_level = level
        family_index = 0
        current_head: str | None = None
        for raw_line in _read_text_lines(fp, encoding=encoding):
            stripped = raw_line.lstrip()
            if not stripped or not stripped[:1].isalpha():
                continue
            word = _first_list_word(stripped)
            if not word:
                continue
            indented = raw_line[:1].isspace()
            if current_head is None or not indented:
                current_head = word
                family_index += 1
                entries += 1

            family_rank = (level - 1) * 1000 + min(family_index, 1000)
            entry = {
                "head": current_head,
                "rank": family_rank,
                "level": level,
                "source_file": os.path.basename(fp),
            }
            if word in rank:
                collisions += 1
                continue
            if word != current_head:
                variants += 1
            rank[word] = entry

    meta = {
        "entries": entries,
        "keys": len(rank),
        "variants": variants,
        "collisions": collisions,
        "max_rank": max_level * 1000,
        "n_levels": max_level,
        "lookup_unit": "range_word_family",
        "source_files": [os.path.basename(fp) for fp in files],
    }
    return rank, meta


def level_of(r):
    return None if r is None else (r - 1) // 1000 + 1


# --------------------------------------------------------------------------- #
# Token → headword / rank mapping
# --------------------------------------------------------------------------- #
def _rank_entry(value, fallback_head: str) -> tuple[str, int | None]:
    if isinstance(value, dict):
        head = str(value.get("head") or fallback_head).lower()
        return head, value.get("rank")
    return fallback_head, value


def _map_tokens_with_diagnostics(tokens, rank_map, lemmatizer):
    """Map tokens and return aggregate diagnostics for the mapping path.

    The four path categories are mutually exclusive. Their token counts therefore
    sum to ``input_tokens``. Diagnostics deliberately contain no submitted terms.
    """
    out = []
    surface_types = set()
    path_counts = Counter()
    for tok in tokens:
        t = tok.lower()
        surface_types.add(t)
        if t in rank_map:
            out.append(_rank_entry(rank_map[t], t))
            path_counts["surface_hit"] += 1
            continue

        normalized = lemmatizer.normalize(t)
        if normalized in rank_map:
            out.append(_rank_entry(rank_map[normalized], normalized))
            path_counts["normalized_fallback_hit"] += 1
        else:
            mapped_unit = normalized or t
            out.append((mapped_unit, OFF_LIST))
            if normalized and normalized != t:
                path_counts["normalized_off_list"] += 1
            else:
                path_counts["identity_fallback"] += 1

    n = len(out)

    def rate(category: str) -> float:
        return path_counts[category] / n if n else 0.0

    mapped_types = {head for head, _rank in out}
    diagnostics = {
        "method_id": PANEL_B_MAPPING_METHOD_ID,
        "input_tokens": n,
        "input_surface_types": len(surface_types),
        "mapped_unit_types": len(mapped_types),
        "collapsed_surface_types": len(surface_types) - len(mapped_types),
        "surface_hit_tokens": path_counts["surface_hit"],
        "surface_hit_rate": rate("surface_hit"),
        "normalized_fallback_hit_tokens": path_counts["normalized_fallback_hit"],
        "normalized_fallback_hit_rate": rate("normalized_fallback_hit"),
        "normalized_off_list_tokens": path_counts["normalized_off_list"],
        "normalized_off_list_rate": rate("normalized_off_list"),
        "identity_fallback_tokens": path_counts["identity_fallback"],
        "identity_fallback_rate": rate("identity_fallback"),
    }
    return out, diagnostics


def map_tokens(tokens, rank_map, lemmatizer):
    """Return list of ``(head, rank)`` per token.

    In-list surface hits are taken at face value and are not re-normalized. Off-list
    tokens are keyed by their normalized fallback head, so two off-list inflections
    sharing a head collapse to one off-list type. Because of the in-list short-circuit,
    Panel A surface-token type counts and Panel B mapped-unit type counts can
    legitimately differ for spellings that the selected list ranks separately.

    Word-family lists store richer rank entries, so an in-list surface hit maps to
    the family head rather than to the literal surface form.
    """
    mapped, _diagnostics = _map_tokens_with_diagnostics(tokens, rank_map, lemmatizer)
    return mapped


# --------------------------------------------------------------------------- #
# Lexical Frequency Profile + coverage
# --------------------------------------------------------------------------- #
def lexical_frequency_profile(mapped, n_levels=8):
    n = len(mapped)
    tok_by_level = Counter()
    types_by_level = defaultdict(set)
    for head, r in mapped:
        lev = level_of(r) or "off"
        tok_by_level[lev] += 1
        types_by_level[lev].add(head)

    rows, cum = [], 0.0
    for lev in list(range(1, n_levels + 1)) + ["off"]:
        toks = tok_by_level.get(lev, 0)
        cov = toks / n if n else 0.0
        if lev != "off":
            cum += cov
        rows.append({
            "level": f"K{lev}" if lev != "off" else "off-list",
            "tokens": toks,
            "types": len(types_by_level.get(lev, set())),
            "coverage_%": round(100 * cov, 2),
            "cumulative_%": round(100 * cum, 2),
        })
    return rows


def coverage_threshold(mapped, thresholds=(90, 95, 98), n_levels=8):
    """For each threshold, the smallest selected-list band K whose cumulative
    matched *text* coverage reaches it (None if it is never reached).

    Coverage = selected-list tokens up to band K / **total** tokens, matching the
    LFP denominator. Unclassified off-list items remain in the denominator and
    cap this selected-list match rate. That is not a claim that such items are
    unknown; lexical-coverage reporting should classify/credit proper nouns,
    marginal words, transparent compounds, acronyms, etc. according to the study's
    stated policy.
    """
    n = len(mapped)
    tok_by_level = Counter(level_of(r) for _, r in mapped if r is not None)
    result = {}
    for thr in thresholds:
        cum = 0.0
        hit = None
        for lev in range(1, n_levels + 1):
            cum += tok_by_level.get(lev, 0) / n if n else 0
            if 100 * cum >= thr:
                hit = lev
                break
        result[thr] = hit  # e.g. 5 -> "need K5"; None -> not reached
    return result


# --------------------------------------------------------------------------- #
# Advanced Guiraud, %beyond-K, mean rank
# --------------------------------------------------------------------------- #
def advanced_guiraud(mapped, advanced_cutoff=2, count_off_list=True):
    """AG = advanced_types / sqrt(N). 'Advanced' = level > cutoff (and off-list
    if count_off_list). Daller, van Hout & Treffers-Daller (2003)."""
    n = len(mapped)
    if n == 0:
        return IDX.NAN
    advanced_types = set()
    for head, r in mapped:
        if r is None:
            if count_off_list:
                advanced_types.add(head)
        elif level_of(r) > advanced_cutoff:
            advanced_types.add(head)
    return len(advanced_types) / math.sqrt(n)


def pct_beyond_k(mapped, k=2):
    """Beyond-K % (Laufer 'Beyond 2000', Kojima & Yamashita 2014 Eq.1): distinct
    advanced word FAMILIES / distinct total FAMILIES × 100. Counting unit is the
    family (here approximated by the normalized head), NOT tokens."""
    if not mapped:
        return IDX.NAN
    fam_total = {head for head, _ in mapped}
    fam_beyond = {head for head, r in mapped if r is None or level_of(r) > k}
    return 100 * len(fam_beyond) / len(fam_total)


def mean_rank(mapped):
    in_list = [r for _, r in mapped if r is not None]
    n = len(mapped)
    if not in_list:
        return {"mean_rank": IDX.NAN, "mean_log_rank": IDX.NAN, "pct_off_list": 100.0}
    return {
        "mean_rank": sum(in_list) / len(in_list),
        "mean_log_rank": sum(math.log(r) for r in in_list) / len(in_list),
        "pct_off_list": 100 * (n - len(in_list)) / n,
    }


# --------------------------------------------------------------------------- #
# P_Lex and S
# --------------------------------------------------------------------------- #
def p_lex(mapped, segment=10, difficult_above_level=1):
    """P_Lex λ (Meara & Bell 2001).

    Split into consecutive 10-word segments; count 'hard' words per segment (hard
    = beyond the first 1000 / off-list); fit the distribution of per-segment counts
    to a Poisson PMF P(k)=λ^k e^-λ/k! by least squares and return the best-fit λ
    (higher λ = richer). λ typically ~0-4.5.

    Caveat: Meara & Bell treat proper nouns/numbers as 'easy'; here proper nouns
    absent from the list count as hard, which can slightly inflate λ.
    """
    segs = [mapped[i:i + segment] for i in range(0, len(mapped), segment)
            if len(mapped[i:i + segment]) == segment]
    if not segs:
        return {"lambda": IDX.NAN, "n_segments": 0}

    def hard(pair):
        _, r = pair
        return r is None or level_of(r) > difficult_above_level

    counts = [sum(hard(p) for p in seg) for seg in segs]
    K = len(segs)
    obs = [counts.count(k) / K for k in range(segment + 1)]

    def sse(lam):
        return sum((obs[k] - math.exp(-lam) * lam ** k / math.factorial(k)) ** 2
                   for k in range(segment + 1))

    # λ is bounded by the segment length: at most every token is hard.
    grid = [0.01 * i for i in range(0, 100 * segment + 1)]   # 0.00 .. segment
    best = min(grid, key=sse)
    fine = [x for x in (best - 0.01 + 0.0005 * i for i in range(41)) if 0 <= x <= segment]
    best = min([best, *fine], key=sse)
    fit = [math.exp(-best) * best ** k / math.factorial(k) for k in range(segment + 1)]
    return {
        "lambda": best,
        "n_segments": K,
        "mean_hard_per_seg": sum(counts) / K,
        "observed_distribution": {k: obs[k] for k in range(segment + 1)},
        "fitted_distribution": {k: fit[k] for k in range(segment + 1)},
    }


def s_index(mapped, sample=50, ranks=(500, 1000, 1500, 2000, 2500, 3000),
            s_max=30000):
    """S (Kojima & Yamashita 2014, Eq.4): fit the coverage curve

        C(x) = (ln x / ln S) * 100

    to the text's empirical cumulative coverage at frequency ranks x, returning the
    fitted S — the rank at which coverage is expected to reach 100 %.

    Procedure (K&Y pp.28): slide a 50-token window (step 1, wrap-around at the end);
    at each window record cumulative coverage at ranks 500..3000; average over
    windows; least-squares fit S.

    NOTE: K&Y use the BNC-spoken family lists; here we use the selected list's ranks,
    so S is method-faithful but NOT numerically comparable to K&Y's published S
    values (~2000-3500).
    """
    n = len(mapped)
    if n < sample:
        return {"S": None, "note": "n < sample (50)"}
    seq = [r for _, r in mapped]
    extended = seq + seq[:sample]                      # wrap-around at text end
    cov = {x: 0.0 for x in ranks}
    for i in range(n):
        win = extended[i:i + sample]
        for x in ranks:
            cov[x] += sum(1 for r in win if r is not None and r <= x) / sample
    emp = {x: 100 * cov[x] / n for x in ranks}         # empirical C(x) in %

    def sse(s):
        ln_s = math.log(s)
        return sum((emp[x] - (math.log(x) / ln_s) * 100) ** 2 for x in ranks)

    # S can be below or above the sampled ranks (K&Y mean S ~ 2285): search wide.
    lo_s = 200
    best = min(range(lo_s, s_max + 1, 50), key=sse)    # coarse…
    best = min(range(max(lo_s, best - 50), min(s_max + 1, best + 51)), key=sse)  # …refine
    capped = best >= s_max
    return {"S": float(min(best, s_max)),
            "empirical_coverage_pct": {x: round(emp[x], 2) for x in ranks},
            "capped": capped,
            "reference_list_note": (
                f"uses the selected list's ranks, not K&Y's "
                f"BNC-spoken family lists; 'capped' = coverage never approaches "
                f"100% within rank {max(ranks)}, so S is not interpretable "
                f"(typical when the selected list leaves substantial text off-list)"
            )}


# --------------------------------------------------------------------------- #
# Band-wise diversity with short-band warnings
# --------------------------------------------------------------------------- #
def band_wise_diversity(mapped, tokens, n_levels=8, min_tokens=50,
                        index_keys=("mtld", "mattr", "hdd"),
                        mtld_threshold=0.72, mattr_window=50, hdd_sample=42):
    """For each band, raw token/type counts (always) plus selected diversity
    indices. Advisory floors are reported separately in ``Min N``. Requested
    windows and samples are never reduced to the band size, so a method that is
    outside its computational domain returns a missing value.

    ``tokens`` is the original token list aligned 1:1 with ``mapped`` so each
    band's words keep their order (needed by MTLD/MATTR).
    """
    by_level_tokens = defaultdict(list)
    by_level_heads = defaultdict(list)
    for (head, r), tok in zip(mapped, tokens):
        lev = level_of(r) or "off"
        by_level_tokens[lev].append(tok)
        by_level_heads[lev].append(head)

    rows = []
    for lev in list(range(1, n_levels + 1)) + ["off"]:
        heads = by_level_heads.get(lev, [])
        nb = len(heads)
        max_floor = 0
        row = {"level": f"K{lev}" if lev != "off" else "off-list",
               "tokens": nb, "types": len(set(heads))}
        for key in index_keys:
            floor = IDX.effective_min_tokens(
                key,
                mtld_threshold=mtld_threshold,
                window=mattr_window,
                hdd_sample=hdd_sample,
                min_tokens_override=min_tokens,
            )
            max_floor = max(max_floor, floor)
            if key == "mtld":
                row[IDX.PRETTY[key]] = IDX.mtld(heads, threshold=mtld_threshold)
            elif key == "mattr":
                row[IDX.PRETTY[key]] = IDX.mattr(heads, window=int(mattr_window))
            elif key == "hdd":
                row[IDX.PRETTY[key]] = IDX.hdd(heads, sample=int(hdd_sample))
            else:
                row[IDX.PRETTY[key]] = IDX._FUNCS[key](heads)
        row["Min N"] = max_floor
        rows.append(row)
    return rows


# --------------------------------------------------------------------------- #
# Convenience: full Panel B
# --------------------------------------------------------------------------- #
def panel_b(tokens, rank_map, lemmatizer, *, n_levels=8, thresholds=(90, 95, 98),
            advanced_cutoff=2, min_tokens=50, mtld_threshold=0.72,
            mattr_window=50, hdd_sample=42):
    mapped, mapping_diagnostics = _map_tokens_with_diagnostics(
        tokens, rank_map, lemmatizer
    )
    return {
        "mapping_diagnostics": mapping_diagnostics,
        "lfp": lexical_frequency_profile(mapped, n_levels),
        "coverage_threshold": coverage_threshold(mapped, thresholds, n_levels),
        "advanced_guiraud": advanced_guiraud(mapped, advanced_cutoff),
        "pct_beyond_k": pct_beyond_k(mapped, advanced_cutoff),
        "mean_rank": mean_rank(mapped),
        "p_lex": p_lex(mapped),
        "s_index": s_index(mapped),
        "band_wise": band_wise_diversity(
            mapped,
            tokens,
            n_levels,
            min_tokens,
            mtld_threshold=mtld_threshold,
            mattr_window=mattr_window,
            hdd_sample=hdd_sample,
        ),
        "_mapped": mapped,
    }
